"""
Import service for uploading Excel files (clients, quotes, purchase orders).

Behavior:
- Mode "replace" (default for restoring backups): wipes existing records and
  inserts everything from the file. Records keep their imported date so that
  list views remain ordered most-recent-first.
- Mode "append": only inserts new records (duplicates skipped).
"""
import uuid
import io
import re
from datetime import datetime, timezone
from fastapi import UploadFile, HTTPException
import openpyxl


# ---------------------- helpers ----------------------

def _norm(s) -> str:
    if s is None:
        return ""
    return str(s).strip().lower()


def _parse_date(val) -> str | None:
    """Return ISO-8601 timestamp from an Excel cell value. Falls back to None.
    Accepts datetime objects (openpyxl native) and strings like 'd/m/yyyy' or
    'yyyy-mm-dd'.
    """
    if val is None or val == "":
        return None
    if isinstance(val, datetime):
        if val.tzinfo is None:
            val = val.replace(tzinfo=timezone.utc)
        return val.isoformat()
    s = str(val).strip()
    # ISO already
    try:
        if "T" in s or re.match(r"^\d{4}-\d{2}-\d{2}", s):
            d = datetime.fromisoformat(s.replace("Z", "+00:00"))
            if d.tzinfo is None:
                d = d.replace(tzinfo=timezone.utc)
            return d.isoformat()
    except Exception:
        pass
    # d/m/yyyy or m/d/yyyy
    m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{2,4})$", s)
    if m:
        a, b, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if y < 100:
            y += 2000
        # Heuristic: Latin American format is d/m/yyyy
        try:
            return datetime(y, b, a, tzinfo=timezone.utc).isoformat()
        except ValueError:
            try:
                return datetime(y, a, b, tzinfo=timezone.utc).isoformat()
            except ValueError:
                return None
    return None


def _read_sheet(file_bytes: bytes):
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"No se pudo leer el archivo Excel: {e}")
    ws = wb.active
    raw_headers = [cell.value for cell in ws[1]]
    return ws, raw_headers


# ---------------------- clients ----------------------

CLIENTS_HEADER_MAP = {
    # Spanish exports
    "empresa / nombre": "name",
    "nombre": "name",
    "empresa": "name",
    "contacto": "contact_person",
    "email": "email",
    "correo": "email",
    "email comercial": "commercial_email",
    "correo comercial": "commercial_email",
    "teléfono": "phone",
    "telefono": "phone",
    "celular": "phone",
    "ciudad": "city",
    "dirección": "address",
    "direccion": "address",
    "ruc / ci": "tax_id",
    "ruc": "tax_id",
    "ci": "tax_id",
    "sector": "sector",
    "notas": "notes",
    "creado": "created_at",
    "fecha": "created_at",
    "fecha de creación": "created_at",
    "fecha de creacion": "created_at",
}


async def import_clients(file: UploadFile, db, mode: str = "replace") -> dict:
    """Import clients from xlsx.

    mode='replace' (default): wipe existing active clients and insert from file.
    mode='append': insert only new (duplicates by email/phone are skipped).
    """
    contents = await file.read()
    ws, raw_headers = _read_sheet(contents)

    col_map = {}
    for i, h in enumerate(raw_headers):
        key = _norm(h)
        if key in CLIENTS_HEADER_MAP:
            col_map[i] = CLIENTS_HEADER_MAP[key]

    if not any(v == "name" for v in col_map.values()):
        raise HTTPException(
            status_code=400,
            detail=(
                "El archivo no tiene la columna obligatoria 'Empresa / Nombre' "
                "(o 'Nombre' / 'Empresa'). Encabezados detectados: "
                f"{[h for h in raw_headers if h]}"
            ),
        )

    # Collect rows
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        data = {}
        for i, val in enumerate(row):
            if i in col_map and val is not None:
                target = col_map[i]
                if target == "created_at":
                    parsed = _parse_date(val)
                    if parsed:
                        data[target] = parsed
                else:
                    data[target] = str(val).strip() if not isinstance(val, str) else val.strip()
        if data.get("name"):
            rows.append(data)

    if not rows:
        return {"inserted": 0, "skipped": 0, "deleted": 0, "mode": mode,
                "message": "No se encontraron clientes válidos en el archivo."}

    deleted = 0
    if mode == "replace":
        # Soft-archive existing active clients to be safe (preserves history),
        # then insert the new set so the listing reflects ONLY the imported file.
        result = await db.clients.update_many(
            {"is_deleted": False},
            {"$set": {"is_deleted": True, "deleted_at": datetime.now(timezone.utc).isoformat(),
                       "deleted_by": "import_replace"}},
        )
        deleted = result.modified_count

    inserted = 0
    skipped = 0
    now_iso = datetime.now(timezone.utc).isoformat()

    for data in rows:
        email = data.get("email", "")
        phone = data.get("phone", "")

        if mode == "append":
            is_dup = False
            if email:
                if await db.clients.find_one({"email": email, "is_deleted": False}, {"_id": 1}):
                    is_dup = True
            if not is_dup and phone:
                clean_phone = re.sub(r"[\s\-\+]", "", phone)
                if len(clean_phone) >= 8 and await db.clients.find_one(
                    {"phone": {"$regex": clean_phone[-8:]}, "is_deleted": False}, {"_id": 1}
                ):
                    is_dup = True
            if is_dup:
                skipped += 1
                continue

        client = {
            "id": str(uuid.uuid4()),
            "name": data.get("name", ""),
            "email": email,
            "commercial_email": data.get("commercial_email", ""),
            "phone": phone,
            "contact_person": data.get("contact_person", ""),
            "address": data.get("address", ""),
            "city": data.get("city", ""),
            "tax_id": data.get("tax_id", ""),
            "sector": data.get("sector", ""),
            "sector_details": "",
            "notes": data.get("notes", ""),
            "is_deleted": False,
            "deleted_at": None,
            "created_at": data.get("created_at") or now_iso,
            "source": "manual",
        }
        await db.clients.insert_one(client)
        inserted += 1

    return {"inserted": inserted, "skipped": skipped, "deleted": deleted, "mode": mode}


# ---------------------- quotes / POs ----------------------

QUOTES_HEADER_MAP = {
    "número": "quote_number",
    "numero": "quote_number",
    "nro": "quote_number",
    "no. cotización": "quote_number",
    "no. orden": "quote_number",
    "factura": "factura",
    "cliente": "client_name",
    "contacto": "client_contact",
    "email cliente": "client_email",
    "email": "client_email",
    "producto": "product_name",
    "código": "product_code",
    "codigo": "product_code",
    "descripción": "description",
    "descripcion": "description",
    "cantidad": "quantity",
    "cant": "quantity",
    "precio unitario": "unit_price",
    "precio": "unit_price",
    "total": "total_price",
    "condiciones de pago": "payment_terms",
    "pago": "payment_terms",
    "validez": "validity",
    "tiempo de entrega": "delivery_time",
    "entrega": "delivery_time",
    "estado": "status",
    "creado": "created_at",
    "fecha": "created_at",
}


async def import_quotes(file: UploadFile, db, doc_type: str = "QUOTE",
                         user: dict = None, mode: str = "replace") -> dict:
    """Import quotes/POs from xlsx.

    mode='replace' (default): delete all existing quotes/POs of this doc_type,
    then insert from file. Preserves the file's original date order in the DB
    so that the listing (sorted by created_at DESC) shows the most recent first.
    mode='append': only inserts quotes whose quote_number is new.
    """
    contents = await file.read()
    ws, raw_headers = _read_sheet(contents)

    col_map = {}
    for i, h in enumerate(raw_headers):
        key = _norm(h)
        if key in QUOTES_HEADER_MAP:
            col_map[i] = QUOTES_HEADER_MAP[key]

    if not any(v == "quote_number" for v in col_map.values()):
        raise HTTPException(
            status_code=400,
            detail=(
                "El archivo no tiene la columna obligatoria 'Número'. "
                f"Encabezados detectados: {[h for h in raw_headers if h]}"
            ),
        )

    # Group rows by quote_number, preserving file order
    quotes_data: dict[str, dict] = {}
    order: list[str] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        data = {}
        for i, val in enumerate(row):
            if i in col_map and val is not None:
                target = col_map[i]
                if target == "created_at":
                    parsed = _parse_date(val)
                    if parsed:
                        data[target] = parsed
                else:
                    if isinstance(val, (int, float)):
                        data[target] = val
                    else:
                        data[target] = str(val).strip()

        qnum = str(data.get("quote_number", "")).strip()
        if not qnum:
            continue

        if qnum not in quotes_data:
            quotes_data[qnum] = {"info": data, "items": []}
            order.append(qnum)
        else:
            # Update created_at if a row has it and the group didn't yet
            if data.get("created_at") and not quotes_data[qnum]["info"].get("created_at"):
                quotes_data[qnum]["info"]["created_at"] = data["created_at"]

        if data.get("product_name") or data.get("product_code"):
            try:
                qty = float(data.get("quantity", 0) or 0)
                price = float(data.get("unit_price", 0) or 0)
            except (ValueError, TypeError):
                qty = 0.0
                price = 0.0
            quotes_data[qnum]["items"].append({
                "item_id": str(uuid.uuid4()),
                "product_id": "",
                "code": str(data.get("product_code", "")),
                "name": str(data.get("product_name", "")),
                "description": str(data.get("description", "")),
                "quantity": qty,
                "unit_price": price,
                "total_price": qty * price,
                "image_url": "",  # populated below from inventory lookup
                "categories": [],
                "discount_amount": 0,
                "discount_type": "none",
                "additional_amount": 0,
                "additional_type": "none",
                "otros": "",
            })

    if not quotes_data:
        return {"inserted": 0, "skipped": 0, "deleted": 0, "mode": mode,
                "message": "No se encontraron cotizaciones/órdenes válidas en el archivo."}

    # Build code -> image_url map from inventory so imported items render
    # thumbnails in PDFs (PDF generator reads item.image_url).
    code_to_image: dict[str, str] = {}
    needed_codes = {it["code"] for q in quotes_data.values() for it in q["items"] if it.get("code")}
    if needed_codes:
        async for prod in db.products.find(
            {"code": {"$in": list(needed_codes)}},
            {"_id": 0, "code": 1, "image_url": 1}
        ):
            url = (prod.get("image_url") or "").strip()
            if url:
                code_to_image[prod["code"]] = url

    deleted = 0
    if mode == "replace":
        result = await db.quotes_v2.update_many(
            {"doc_type": doc_type, "is_deleted": False},
            {"$set": {"is_deleted": True, "deleted_at": datetime.now(timezone.utc).isoformat(),
                       "deleted_by": "import_replace"}},
        )
        deleted = result.modified_count

    inserted = 0
    skipped = 0
    now = datetime.now(timezone.utc)

    # Preserve file order: assign created_at as the parsed date when available;
    # otherwise stagger by 1 second descending so the FIRST row in the file
    # ends up as the MOST recent in the listing (matches user expectation).
    for idx, qnum in enumerate(order):
        qdata = quotes_data[qnum]
        info = qdata["info"]
        items = qdata["items"]

        if mode == "append":
            existing = await db.quotes_v2.find_one(
                {"quote_number": qnum, "doc_type": doc_type, "is_deleted": False},
                {"_id": 1},
            )
            if existing:
                skipped += 1
                continue

        subtotal = sum(it["total_price"] for it in items)
        tax = subtotal * 0.15
        total = subtotal + tax

        # Hydrate items with the inventory's image_url so PDFs render thumbs
        for it in items:
            if not it.get("image_url") and it.get("code"):
                it["image_url"] = code_to_image.get(it["code"], "")

        created_at = info.get("created_at")
        if not created_at:
            # Stagger so the file order is preserved in DESC listing
            from datetime import timedelta as _td
            created_at = (now - _td(seconds=idx)).isoformat()

        quote = {
            "id": str(uuid.uuid4()),
            "doc_type": doc_type,
            "quote_number": qnum,
            "client_id": "",
            "client_name": str(info.get("client_name", "")),
            "client_contact": str(info.get("client_contact", "")),
            "client_email": str(info.get("client_email", "")),
            "items": items,
            "subtotal": subtotal,
            "tax": tax,
            "total": total,
            "status": str(info.get("status", "draft")).lower() if info.get("status") else "draft",
            "payment_terms": str(info.get("payment_terms", "")),
            "validity": str(info.get("validity", "15 días")),
            "delivery_time": str(info.get("delivery_time", "")),
            "factura": str(info.get("factura", "")) if info.get("factura") else "",
            "is_deleted": False,
            "deleted_at": None,
            "created_by_id": user.get("id", "") if user else "",
            "created_by_name": user.get("name", "") if user else "",
            "created_at": created_at,
        }
        await db.quotes_v2.insert_one(quote)
        inserted += 1

    return {"inserted": inserted, "skipped": skipped, "deleted": deleted, "mode": mode}
